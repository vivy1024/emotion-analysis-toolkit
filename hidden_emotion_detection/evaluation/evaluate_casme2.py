#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
CASME II 离线评估脚本
在 CASME II 数据集上测试光流 Spotting 引擎，支持 LOSO 评估协议和 STRS 指标。

用法:
    python -m hidden_emotion_detection.evaluation.evaluate_casme2 \
        --data_root "旧有文件/18/data/raw/CASME II" \
        --predictor "旧有文件/18/data/raw/CASME II/shape_predictor_68_face_landmarks.dat"

可选参数:
    --subjects 01 02 03    # 只评估指定被试（调试用）
    --peak_threshold 0.5   # 光流峰值阈值
    --output results.csv   # 输出预测结果
"""

import argparse
import logging
import os
import sys
import json
import time
import numpy as np
import cv2
from pathlib import Path
from typing import Dict, List, Tuple, Optional

# 配置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(name)s: %(message)s'
)
logger = logging.getLogger("CASME2_Eval")


# ============================================================
# 数据加载
# ============================================================

def load_casme2_annotations(data_root: str) -> List[dict]:
    """
    加载 CASME II 标注文件

    Returns:
        [{'subject': '01', 'video': 'EP02_01f', 'onset': 46, 'apex': 59,
          'offset': 86, 'au': '12', 'emotion': 'happiness', 'obj_class': 1}, ...]
    """
    import openpyxl

    # 优先使用 CASME2_Metadata.xlsx（包含所有信息）
    metadata_path = os.path.join(data_root, 'CASME2_Metadata.xlsx')
    if not os.path.exists(metadata_path):
        # 备选：excel/CASME2.xlsx
        metadata_path = os.path.join(data_root, 'excel', 'CASME2.xlsx')

    if not os.path.exists(metadata_path):
        raise FileNotFoundError(f"找不到标注文件: {metadata_path}")

    wb = openpyxl.load_workbook(metadata_path, read_only=True)
    ws = wb.active

    annotations = []
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue

        subject = str(row[0]).zfill(2)
        video = str(row[1])

        # 安全解析整数（有些字段可能是 '/' 等非数字）
        def safe_int(val, default=0):
            if val is None:
                return default
            try:
                return int(val)
            except (ValueError, TypeError):
                return default

        onset = safe_int(row[2])
        apex = safe_int(row[3])
        offset = safe_int(row[4])

        # AU 和 Emotion 列位置取决于文件格式
        au = str(row[5]) if row[5] is not None else ''
        emotion = str(row[6]).lower() if row[6] is not None else 'others'

        # Objective Class（如果有）
        obj_class = int(row[8]) if len(row) > 8 and row[8] is not None else 0

        annotations.append({
            'subject': subject,
            'video': video,
            'onset': onset,
            'apex': apex,
            'offset': offset,
            'au': au,
            'emotion': emotion,
            'obj_class': obj_class,
        })

    wb.close()
    logger.info(f"加载 {len(annotations)} 条标注 (来自 {metadata_path})")
    return annotations


def _imread_unicode(path: str, flags=cv2.IMREAD_GRAYSCALE) -> Optional[np.ndarray]:
    """支持中文路径的 imread"""
    try:
        # 先尝试直接读取
        img = cv2.imread(path, flags)
        if img is not None:
            return img
    except Exception:
        pass
    # 用 numpy 中转（处理中文路径）
    try:
        with open(path, 'rb') as f:
            data = np.frombuffer(f.read(), dtype=np.uint8)
        return cv2.imdecode(data, flags)
    except Exception:
        return None


def load_video_frames(data_root: str, subject: str, video: str) -> List[Tuple[int, np.ndarray]]:
    """
    加载一个视频的所有帧（CASME II 以图片序列存储）

    Returns:
        [(frame_id, gray_image), ...] 按帧号排序
    """
    video_dir = os.path.join(data_root, f'sub{subject}', video)
    if not os.path.isdir(video_dir):
        return []

    frames = []
    for fname in os.listdir(video_dir):
        if not fname.lower().endswith(('.jpg', '.png', '.bmp')):
            continue
        # 提取帧号: img31.jpg -> 31
        try:
            frame_id = int(''.join(filter(str.isdigit, os.path.splitext(fname)[0])))
        except ValueError:
            continue

        img_path = os.path.join(video_dir, fname)
        img = _imread_unicode(img_path)
        if img is not None:
            frames.append((frame_id, img))

    frames.sort(key=lambda x: x[0])
    return frames


# ============================================================
# 光流 Spotting 核心算法（离线版，不依赖 EventBus）
# ============================================================

class OfflineOpticalFlowSpotter:
    """离线光流 Spotting 检测器"""

    # ROI 区域定义
    ROI_DEFINITIONS = {
        'left_eyebrow': {'landmarks': list(range(17, 22)), 'percent': 0.3},
        'right_eyebrow': {'landmarks': list(range(22, 27)), 'percent': 0.3},
        'left_eye': {'landmarks': list(range(36, 42)), 'percent': 0.2},
        'right_eye': {'landmarks': list(range(42, 48)), 'percent': 0.2},
        'nose': {'landmarks': list(range(27, 36)), 'percent': 0.2},
        'mouth_outer': {'landmarks': list(range(48, 60)), 'percent': 0.3},
        'mouth_inner': {'landmarks': list(range(60, 68)), 'percent': 0.3},
    }
    GLOBAL_REF_LANDMARKS = [29, 30, 31]

    def __init__(self, predictor_path: str, config: dict = None):
        """
        Args:
            predictor_path: dlib shape_predictor_68_face_landmarks.dat 路径
            config: 配置参数
        """
        import dlib
        self.detector = dlib.get_frontal_face_detector()
        # dlib 不支持中文路径，需要确保路径是 ASCII 兼容的
        predictor_path_str = str(predictor_path)
        try:
            predictor_path_str.encode('ascii')
        except UnicodeEncodeError:
            # 中文路径：复制到临时目录
            import tempfile, shutil
            tmp_path = os.path.join(tempfile.gettempdir(), 'shape_predictor_68.dat')
            if not os.path.exists(tmp_path):
                shutil.copy2(predictor_path_str, tmp_path)
            predictor_path_str = tmp_path
        self.predictor = dlib.shape_predictor(predictor_path_str)
        self.config = config or {}

        # 光流参数
        self.flow_params = dict(
            pyr_scale=0.5, levels=3, winsize=15,
            iterations=5, poly_n=7, poly_sigma=1.5, flags=0
        )

        # Spotting 阈值
        self.peak_threshold = self.config.get('peak_threshold', 0.1)
        self.valley_ratio = self.config.get('valley_ratio', 0.33)
        self.min_duration = self.config.get('min_duration_frames', 3)
        self.max_duration = self.config.get('max_duration_frames', 90)
        self.merge_gap = self.config.get('merge_gap', 3)
        self.nms_iou = self.config.get('nms_iou_threshold', 0.3)

    def detect_landmarks(self, gray: np.ndarray) -> Optional[np.ndarray]:
        """检测人脸并返回 68 点关键点"""
        faces = self.detector(gray, 0)
        if len(faces) == 0:
            # 尝试直接用整张图作为人脸区域
            import dlib
            h, w = gray.shape
            faces = [dlib.rectangle(0, 0, w, h)]

        shape = self.predictor(gray, faces[0])
        landmarks = np.array([[shape.part(i).x, shape.part(i).y]
                              for i in range(68)], dtype=np.float32)
        return landmarks

    def process_video(self, frames: List[Tuple[int, np.ndarray]],
                      fps: float = 30.0) -> List[dict]:
        """
        处理一个视频的所有帧，返回检测到的微表情段

        Args:
            frames: [(frame_id, gray_image), ...]
            fps: 帧率

        Returns:
            [{'onset': frame_id, 'offset': frame_id, 'intensity': float, 'type': str}, ...]
        """
        if len(frames) < 5:
            return []

        # 1. 检测所有帧的关键点
        frame_ids = []
        grays = []
        landmarks_list = []

        for frame_id, gray in frames:
            lm = self.detect_landmarks(gray)
            if lm is not None:
                frame_ids.append(frame_id)
                grays.append(gray)
                landmarks_list.append(lm)

        if len(grays) < 5:
            return []

        # 2. 帧对齐
        aligned = self._align_frames(grays, landmarks_list)
        if aligned is None:
            return []

        # 3. 计算多区域光流
        roi_flows = self._compute_roi_flows(aligned, landmarks_list)
        if not roi_flows:
            return []

        # 4. 融合信号
        fused = self._fuse_signals(roi_flows)

        # 5. 检测段
        segments = self._detect_segments(fused)

        # 6. NMS
        segments = self._nms(segments)

        # 7. 转换为帧号
        results = []
        for onset_idx, offset_idx, intensity in segments:
            if onset_idx >= len(frame_ids) or offset_idx >= len(frame_ids):
                continue
            onset_fid = frame_ids[onset_idx]
            offset_fid = frame_ids[min(offset_idx, len(frame_ids) - 1)]
            duration_sec = (offset_idx - onset_idx) / fps
            expr_type = 'me' if duration_sec <= 0.5 else 'mae'

            results.append({
                'onset': onset_fid,
                'offset': offset_fid,
                'intensity': float(intensity),
                'type': expr_type,
                'duration_sec': duration_sec,
            })

        return results

    def _align_frames(self, grays, landmarks_list):
        """帧对齐"""
        if len(grays) < 2:
            return None
        ref_lm = landmarks_list[0]
        src_pts = np.float32([ref_lm[39], ref_lm[42], ref_lm[33]])
        aligned = [grays[0]]
        h, w = grays[0].shape[:2]

        for i in range(1, len(grays)):
            lm = landmarks_list[i]
            dst_pts = np.float32([lm[39], lm[42], lm[33]])
            try:
                M = cv2.getAffineTransform(dst_pts, src_pts)
                warped = cv2.warpAffine(grays[i], M, (w, h),
                                        flags=cv2.INTER_LINEAR,
                                        borderMode=cv2.BORDER_REFLECT_101)
                aligned.append(warped)
            except cv2.error:
                aligned.append(grays[i])
        return aligned

    def _compute_roi_flows(self, aligned, landmarks_list):
        """计算各 ROI 光流"""
        n = len(aligned)
        if n < 2:
            return {}

        roi_flows = {name: np.zeros(n - 1) for name in self.ROI_DEFINITIONS}
        global_flows = np.zeros((n - 1, 2))

        for i in range(n - 1):
            flow = cv2.calcOpticalFlowFarneback(
                aligned[i], aligned[i + 1], None, **self.flow_params
            )
            ref_lm = landmarks_list[min(i, len(landmarks_list) - 1)]

            # 全局运动
            gx, gy = self._get_roi_flow(flow, ref_lm, self.GLOBAL_REF_LANDMARKS, 0.5, 10)
            global_flows[i] = [gx, gy]

            # 各 ROI
            for name, roi_def in self.ROI_DEFINITIONS.items():
                lx, ly = self._get_roi_flow(
                    flow, ref_lm, roi_def['landmarks'], roi_def['percent'], 5
                )
                cx, cy = lx - gx, ly - gy
                roi_flows[name][i] = np.sqrt(cx ** 2 + cy ** 2)

        return roi_flows

    @staticmethod
    def _get_roi_flow(flow, landmarks, indices, percent, expand):
        """提取 ROI 区域光流"""
        h, w = flow.shape[:2]
        pts = landmarks[indices].astype(int)
        x_min = max(0, pts[:, 0].min() - expand)
        x_max = min(w, pts[:, 0].max() + expand)
        y_min = max(0, pts[:, 1].min() - expand)
        y_max = min(h, pts[:, 1].max() + expand)

        if x_max <= x_min or y_max <= y_min:
            return 0.0, 0.0

        roi = flow[y_min:y_max, x_min:x_max]
        fx, fy = roi[:, :, 0].flatten(), roi[:, :, 1].flatten()
        mag = np.sqrt(fx ** 2 + fy ** 2)

        if len(mag) == 0:
            return 0.0, 0.0

        threshold = np.percentile(mag, (1 - percent) * 100)
        mask = mag >= threshold
        if mask.sum() == 0:
            return float(fx.mean()), float(fy.mean())
        return float(fx[mask].mean()), float(fy[mask].mean())

    @staticmethod
    def _fuse_signals(roi_flows):
        """融合多 ROI 信号"""
        signals = list(roi_flows.values())
        if not signals:
            return np.array([])
        # 取所有 ROI 的最大值（使用原始光流幅度）
        fused = np.max(np.stack(signals, axis=0), axis=0)
        return fused

    def _detect_segments(self, signal):
        """
        基于光流信号检测微表情时间段
        策略：找到局部峰值，然后向两侧扩展到谷值
        """
        if len(signal) < self.min_duration:
            return []

        # 平滑信号（3帧滑动平均）
        kernel = np.ones(3) / 3
        smoothed = np.convolve(signal, kernel, mode='same')

        # 计算局部基线（21帧窗口的中位数）
        baseline = np.zeros_like(smoothed)
        half_win = 10
        for i in range(len(smoothed)):
            start = max(0, i - half_win)
            end = min(len(smoothed), i + half_win + 1)
            baseline[i] = np.median(smoothed[start:end])

        # 差分信号
        diff = smoothed - baseline

        # 找局部峰值（比两侧邻居都高的点）
        peaks = []
        for i in range(2, len(diff) - 2):
            if diff[i] > diff[i-1] and diff[i] > diff[i+1] and diff[i] > diff[i-2] and diff[i] > diff[i+2]:
                if diff[i] > np.percentile(diff[diff > 0], 50) if np.any(diff > 0) else 0:
                    peaks.append((i, diff[i]))

        if not peaks:
            return []

        # 从每个峰值向两侧扩展到谷值
        segments = []
        for peak_idx, peak_val in peaks:
            # 向左找谷值
            onset = peak_idx
            for j in range(peak_idx - 1, max(0, peak_idx - 30), -1):
                if diff[j] <= 0 or smoothed[j] <= baseline[j]:
                    onset = j
                    break
                onset = j

            # 向右找谷值
            offset = peak_idx
            for j in range(peak_idx + 1, min(len(diff), peak_idx + 30)):
                if diff[j] <= 0 or smoothed[j] <= baseline[j]:
                    offset = j
                    break
                offset = j

            duration = offset - onset
            if self.min_duration <= duration <= self.max_duration:
                segments.append((onset, offset, float(peak_val)))

        # 合并重叠段
        if segments:
            segments.sort(key=lambda x: x[0])
            merged = [segments[0]]
            for seg in segments[1:]:
                prev = merged[-1]
                if seg[0] <= prev[1] + self.merge_gap:
                    # 合并
                    merged[-1] = (prev[0], max(prev[1], seg[1]), max(prev[2], seg[2]))
                else:
                    merged.append(seg)
            segments = merged

        return segments

    def _nms(self, segments):
        """NMS 去重"""
        if len(segments) <= 1:
            return segments
        segments = sorted(segments, key=lambda x: x[2], reverse=True)
        keep = []
        for seg in segments:
            overlap = False
            for kept in keep:
                s1, e1 = seg[0], seg[1]
                s2, e2 = kept[0], kept[1]
                inter = max(0, min(e1, e2) - max(s1, s2))
                union = max(e1, e2) - min(s1, s2)
                iou = inter / union if union > 0 else 0
                if iou > self.nms_iou:
                    overlap = True
                    break
            if not overlap:
                keep.append(seg)
        return keep


# ============================================================
# LOSO 评估协议
# ============================================================

def run_loso_evaluation(data_root: str, predictor_path: str,
                        annotations: List[dict],
                        config: dict = None,
                        subjects: List[str] = None) -> dict:
    """
    Leave-One-Subject-Out 评估

    Args:
        data_root: CASME II 数据根目录
        predictor_path: dlib 模型路径
        annotations: 标注列表
        config: Spotting 配置
        subjects: 指定被试列表（None=全部）

    Returns:
        评估结果字典
    """
    from .strs_metric import STRSEvaluator

    # 获取所有被试
    all_subjects = sorted(set(a['subject'] for a in annotations))
    if subjects:
        all_subjects = [s for s in all_subjects if s in subjects]

    logger.info(f"LOSO 评估: {len(all_subjects)} 个被试")

    # 初始化 Spotter
    spotter = OfflineOpticalFlowSpotter(predictor_path, config)
    evaluator = STRSEvaluator(iou_threshold=0.5)

    all_predictions = []
    all_ground_truths = []
    per_subject_results = {}

    for si, test_subject in enumerate(all_subjects):
        logger.info(f"[{si + 1}/{len(all_subjects)}] 测试被试: sub{test_subject}")

        # 获取该被试的标注
        subject_annotations = [a for a in annotations if a['subject'] == test_subject]
        # 获取该被试的所有视频
        subject_videos = sorted(set(a['video'] for a in subject_annotations))

        subject_preds = []
        subject_gts = []

        for video_name in subject_videos:
            # 加载视频帧
            frames = load_video_frames(data_root, test_subject, video_name)
            if not frames:
                logger.warning(f"  跳过 sub{test_subject}/{video_name}: 无帧数据")
                continue

            logger.info(f"  处理 {video_name}: {len(frames)} 帧")

            # 运行 Spotting
            t0 = time.time()
            detections = spotter.process_video(frames, fps=30.0)
            elapsed = time.time() - t0
            logger.info(f"  检测到 {len(detections)} 个段 ({elapsed:.1f}s)")

            # 转换为评估格式
            for det in detections:
                subject_preds.append({
                    'subject': test_subject,
                    'video': video_name,
                    'onset': det['onset'],
                    'offset': det['offset'],
                    'emotion': 'unknown',  # Spotting only，暂无分类
                    'intensity': det['intensity'],
                    'type': det['type'],
                })

            # GT
            video_gts = [a for a in subject_annotations if a['video'] == video_name]
            for gt in video_gts:
                subject_gts.append({
                    'subject': test_subject,
                    'video': video_name,
                    'onset': gt['onset'],
                    'offset': gt['offset'],
                    'emotion': gt['emotion'],
                })

        # 该被试的 Spotting 评估
        spot_result = evaluator.evaluate_spotting(subject_preds, subject_gts)
        per_subject_results[test_subject] = {
            'n_gt': len(subject_gts),
            'n_pred': len(subject_preds),
            'tp': spot_result['tp'],
            'fp': spot_result['fp'],
            'fn': spot_result['fn'],
            'precision': spot_result['precision'],
            'recall': spot_result['recall'],
            'f1': spot_result['f1'],
        }
        logger.info(
            f"  sub{test_subject}: GT={len(subject_gts)}, Pred={len(subject_preds)}, "
            f"TP={spot_result['tp']}, FP={spot_result['fp']}, FN={spot_result['fn']}, "
            f"F1={spot_result['f1']:.4f}"
        )

        all_predictions.extend(subject_preds)
        all_ground_truths.extend(subject_gts)

    # 全局评估
    logger.info("=" * 60)
    logger.info("全局 Spotting 评估")
    global_spot = evaluator.evaluate_spotting(all_predictions, all_ground_truths)
    logger.info(
        f"TP={global_spot['tp']}, FP={global_spot['fp']}, FN={global_spot['fn']}"
    )
    logger.info(
        f"Precision={global_spot['precision']:.4f}, "
        f"Recall={global_spot['recall']:.4f}, "
        f"F1={global_spot['f1']:.4f}"
    )

    return {
        'global_spotting': {
            'tp': global_spot['tp'],
            'fp': global_spot['fp'],
            'fn': global_spot['fn'],
            'precision': global_spot['precision'],
            'recall': global_spot['recall'],
            'f1': global_spot['f1'],
        },
        'per_subject': per_subject_results,
        'predictions': all_predictions,
        'ground_truths': all_ground_truths,
        'config': config or {},
    }


# ============================================================
# 主入口
# ============================================================

def main():
    parser = argparse.ArgumentParser(description='CASME II 光流 Spotting 离线评估')
    parser.add_argument('--data_root', type=str,
                        default=r'旧有文件\18\data\raw\CASME II',
                        help='CASME II 数据根目录')
    parser.add_argument('--predictor', type=str,
                        default=r'hidden_emotion_detection\models\shape_predictor_68_face_landmarks.dat',
                        help='dlib 68点模型路径')
    parser.add_argument('--subjects', nargs='*', default=None,
                        help='指定被试（如 01 02 03），默认全部')
    parser.add_argument('--peak_threshold', type=float, default=0.1,
                        help='光流峰值阈值（差分信号）')
    parser.add_argument('--min_duration', type=int, default=3,
                        help='最小持续帧数')
    parser.add_argument('--max_duration', type=int, default=90,
                        help='最大持续帧数')
    parser.add_argument('--output', type=str, default=None,
                        help='输出预测结果 CSV')
    parser.add_argument('--output_json', type=str, default=None,
                        help='输出完整评估结果 JSON')
    args = parser.parse_args()

    # 解析路径（相对于项目根目录）
    project_root = Path(__file__).resolve().parent.parent.parent
    data_root = Path(args.data_root)
    if not data_root.is_absolute():
        data_root = project_root / data_root
    predictor_path = Path(args.predictor)
    if not predictor_path.is_absolute():
        predictor_path = project_root / predictor_path

    logger.info(f"数据目录: {data_root}")
    logger.info(f"dlib 模型: {predictor_path}")

    if not data_root.exists():
        logger.error(f"数据目录不存在: {data_root}")
        sys.exit(1)
    if not predictor_path.exists():
        logger.error(f"dlib 模型不存在: {predictor_path}")
        sys.exit(1)

    # 加载标注
    annotations = load_casme2_annotations(str(data_root))

    # 配置
    config = {
        'peak_threshold': args.peak_threshold,
        'min_duration_frames': args.min_duration,
        'max_duration_frames': args.max_duration,
    }

    # 运行 LOSO 评估
    results = run_loso_evaluation(
        str(data_root), str(predictor_path),
        annotations, config, args.subjects
    )

    # 输出预测结果
    if args.output:
        import csv
        output_path = Path(args.output)
        if not output_path.is_absolute():
            output_path = project_root / output_path
        with open(output_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=['subject', 'video', 'onset', 'offset', 'emotion', 'type', 'intensity'])
            writer.writeheader()
            for pred in results['predictions']:
                writer.writerow({
                    'subject': pred['subject'],
                    'video': pred['video'],
                    'onset': pred['onset'],
                    'offset': pred['offset'],
                    'emotion': pred.get('emotion', 'unknown'),
                    'type': pred.get('type', ''),
                    'intensity': f"{pred.get('intensity', 0):.4f}",
                })
        logger.info(f"预测结果已保存: {output_path}")

    # 输出 JSON
    if args.output_json:
        json_path = Path(args.output_json)
        if not json_path.is_absolute():
            json_path = project_root / json_path
        # 移除不可序列化的字段
        json_results = {
            'global_spotting': results['global_spotting'],
            'per_subject': results['per_subject'],
            'config': results['config'],
            'n_predictions': len(results['predictions']),
            'n_ground_truths': len(results['ground_truths']),
        }
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(json_results, f, indent=2, ensure_ascii=False)
        logger.info(f"评估结果已保存: {json_path}")

    # 打印汇总
    print("\n" + "=" * 60)
    print("CASME II 光流 Spotting 评估结果")
    print("=" * 60)
    gs = results['global_spotting']
    print(f"总 GT: {gs['tp'] + gs['fn']}")
    print(f"总预测: {gs['tp'] + gs['fp']}")
    print(f"TP: {gs['tp']}, FP: {gs['fp']}, FN: {gs['fn']}")
    print(f"Precision: {gs['precision']:.4f}")
    print(f"Recall:    {gs['recall']:.4f}")
    print(f"F1-score:  {gs['f1']:.4f}")
    print("=" * 60)

    # 各被试结果
    print("\n各被试结果:")
    print(f"{'Subject':>8} {'GT':>4} {'Pred':>5} {'TP':>4} {'FP':>4} {'FN':>4} {'F1':>8}")
    print("-" * 45)
    for subj, sr in sorted(results['per_subject'].items()):
        print(f"  sub{subj:>3} {sr['n_gt']:>4} {sr['n_pred']:>5} "
              f"{sr['tp']:>4} {sr['fp']:>4} {sr['fn']:>4} {sr['f1']:>8.4f}")


if __name__ == '__main__':
    main()
